#! /home/codeer/anaconda3/bin/python3.9
import sqlite3 as sq
import json
import openpyxl as xl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

class xlGen(Workbook):
    def __init__(self):
        super().__init__()

        with open("Test.json", 'r') as data:
            self.tkb = json.load(data)
        data.close()


        self.outlineBorder = Border(left=Side(style='medium'),
                              right=Side(style='medium'),
                              top=Side(style='medium'),
                              bottom=Side(style='medium'))

        self.allBorder = Border(left=Side(style='medium'),
                                right=Side(style='medium'),
                                top=Side(style='medium'),
                                bottom=Side(style='medium'),
                                horizontal=Side(style='medium'),
                                vertical=Side(style='medium'))

        self.bottomRightBorder = Border(bottom=Side(style='medium'),
                                        right=Side(style='medium'))

        self.rightBorder = Border(right=Side(style='medium'))

        
        self.ws = self.active
        self.ws.title = "TKBTD-12A1"
        self.format()
        self.insert()

        self.template = False
        self.save("test.xlsx")
            
        
    def insert(self):
        j = 3
        for thu, listed in self.tkb.items():
            i = 1
            for lop, mon in listed.items():
                self.ws.cell(column=j, row=i+2, value=mon[0][0])
                self.ws.cell(column=j, row=i+3, value=mon[0][1])
                self.ws.cell(column=j, row=i+4, value=mon[1][0])
                self.ws.cell(column=j, row=i+5, value=mon[1][1])
                self.ws.cell(column=j, row=i+6, value=mon[2][0])
                self.ws.cell(column=j, row=i+7, value=mon[2][1])
                self.ws.cell(column=j, row=i+8, value=mon[3][0])
                self.ws.cell(column=j, row=i+9, value=mon[3][1])
                self.ws.cell(column=j, row=i+10, value=mon[4][0])
                self.ws.cell(column=j, row=i+11, value=mon[4][1])
                i += 13
            j += 1
        

                
    def format(self):
        danhSach = self.runQuery("SELECT name FROM sqlite_master", receive=True)
        i = 1

        for j in range(3,9):
            self.ws.column_dimensions[get_column_letter(j)].width = 10.5

        for lop in danhSach:
            self.ws.merge_cells("C{}:H{}".format(i,i))
            self.ws.merge_cells("B{}:B{}".format(i, i+1))
            self.ws.merge_cells("B{}:B{}".format(i+2, i+3))
            self.ws.merge_cells("B{}:B{}".format(i+4, i+5))
            self.ws.merge_cells("B{}:B{}".format(i+6, i+7))
            self.ws.merge_cells("B{}:B{}".format(i+8, i+9))
            self.ws.merge_cells("B{}:B{}".format(i+10, i+11))

            self.drawBorder("B{}:B{}".format(i, i+11), self.allBorder)
            self.drawBorder("C{}:H{}".format(i+1, i+1), self.allBorder)

            self.drawBorder("C{}:C{}".format(i+2, i+11), self.rightBorder)
            self.drawBorder("D{}:D{}".format(i+2, i+11), self.rightBorder)
            self.drawBorder("E{}:E{}".format(i+2, i+11), self.rightBorder)
            self.drawBorder("F{}:F{}".format(i+2, i+11), self.rightBorder)
            self.drawBorder("G{}:G{}".format(i+2, i+11), self.rightBorder)
            self.drawBorder("H{}:H{}".format(i+2, i+11), self.rightBorder)

            self.drawBorder("C{}:H{}".format(i+3, i+3), self.bottomRightBorder)
            self.drawBorder("C{}:H{}".format(i+5, i+5), self.bottomRightBorder)
            self.drawBorder("C{}:H{}".format(i+7, i+7), self.bottomRightBorder)
            self.drawBorder("C{}:H{}".format(i+9, i+9), self.bottomRightBorder)
            self.drawBorder("C{}:H{}".format(i+11, i+11), self.bottomRightBorder)

            self.ws.cell(column=2, row=i+2, value="T1")
            self.ws.cell(column=2, row=i+4, value="T2")
            self.ws.cell(column=2, row=i+6, value="T3")
            self.ws.cell(column=2, row=i+8, value="T4")
            self.ws.cell(column=2, row=i+10, value="T5")

            self.ws.cell(column=3, row=i+1, value="Thứ hai")
            self.ws.cell(column=4, row=i+1, value="Thứ ba")
            self.ws.cell(column=5, row=i+1, value="Thứ tư")
            self.ws.cell(column=6, row=i+1, value="Thứ năm")
            self.ws.cell(column=7, row=i+1, value="Thứ sáu")
            self.ws.cell(column=8, row=i+1, value="Thứ bảy")

            cell1 = self.ws.cell(column=3, row=i, value=lop[0])
            cell1.alignment = Alignment(horizontal='center', vertical='center')
            cell1.border = self.outlineBorder

            cell2 = self.ws.cell(column=8, row=i)
            cell2.border = self.outlineBorder

            i += 13

    def drawBorder(self, cell_range, border):
        for row in self.ws[cell_range]:
            for cell in row:
                cell.border= border

    @staticmethod
    def runQuery(sql, data=None, receive=False):
        conn = sq.connect("TKB.db")
        cursor = conn.cursor()

        if data:
            cursor.execute(sql, data)
        else:
            cursor.execute(sql)

        if receive:
            return cursor.fetchall()
        else:
            conn.commit()

        conn.close()


if __name__ == "__main__":
    xlGen = xlGen()
